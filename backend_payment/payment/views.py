import datetime
import json

import random
import uuid
from http import HTTPStatus
from pprint import pprint
import pandas as pd
import requests
import structlog
from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import PermissionRequiredMixin, AccessMixin, LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import F, Avg, Sum, Count, Window, Q
from django.db.models.functions import TruncDate, ExtractHour
from django.db.transaction import atomic

from django.http import HttpResponse, QueryDict, HttpResponseNotAllowed, HttpResponseForbidden, HttpResponseBadRequest, \
    JsonResponse, HttpResponseRedirect, Http404
from django.shortcuts import render, redirect, get_object_or_404
from django.template.response import TemplateResponse
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.http import urlencode
from django.views.generic import CreateView, DetailView, FormView, UpdateView, ListView, DeleteView
from django_better_admin_arrayfield.admin.mixins import DynamicArrayMixin
from django_currentuser.middleware import get_current_user, get_current_authenticated_user
from openpyxl.workbook import Workbook
from rest_framework.permissions import AllowAny
from rest_framework.views import APIView
from urllib3 import Retry, PoolManager

from core.global_func import hash_gen, TZ, export_payments_func
from deposit.models import Incoming
from payment import forms
from payment.filters import PaymentFilter, WithdrawFilter, BalanceChangeFilter, PaymentMerchStatFilter, \
    MerchPaymentFilter, BalanceFilter
from payment.forms import InvoiceForm, PaymentListConfirmForm, PaymentForm, InvoiceM10Form, InvoiceTestForm, \
     MerchantForm, WithdrawForm, DateFilterForm, InvoiceM10SmsForm, MerchBalanceChangeForm, SupportOptionsForm
from payment.func import work_calc
from payment.models import Payment, PayRequisite, Merchant, PhoneScript, Bank, Withdraw, BalanceChange, Work
from payment.permissions import AuthorRequiredMixin, StaffOnlyPerm, MerchantOnlyPerm, SuperuserOnlyPerm, \
    SupportOrSuperuserPerm, MerchantOrViewPerm
from payment.task import send_payment_webhook, send_withdraw_webhook
from users.models import SupportOptions



logger = structlog.get_logger(__name__)
User = get_user_model()


def make_page_obj(request, objects, numbers_of_posts=settings.PAGINATE):
    paginator = Paginator(objects, numbers_of_posts)
    page_number = request.GET.get('page')
    return paginator.get_page(page_number)


@login_required()
def menu(request, *args, **kwargs):
    template = 'payment/menu.html'
    user = request.user
    context = {}
    if user.is_authenticated:
        merchants = Merchant.objects.filter(owner=user)
        context = {'merchants': merchants}
    return render(request, template_name=template, context=context)


class SupportOptionsView(SupportOrSuperuserPerm, FormView, UpdateView,):
    form_class = SupportOptionsForm
    template_name = 'options.html'
    model = SupportOptions

    def get_object(self, queryset=None):
        obj = SupportOptions.load()
        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        filtered_payments = Payment.objects.filter(pay_type='card_2').filter(status=9).filter(
            confirmed_time__gt=timezone.now() - datetime.timedelta(days=1))
        
        all_steps = (
            filtered_payments
            .annotate(date1=TruncDate('confirmed_time', tzinfo=TZ))
            .annotate(username=F('confirmed_user__username'))
            # .annotate(hour=ExtractHour('confirmed_time'))
            .annotate(step_sum=Window(expression=Sum('confirmed_amount'), partition_by=['confirmed_user']))
            .annotate(step_count=Window(expression=Count('confirmed_amount'), partition_by=['confirmed_user']))
        )

        last_day = all_steps.values('username', 'step_sum', 'step_count').distinct('username').order_by('username')
        context['last_day'] = last_day

        # stat_date = datetime.datetime(2024, 9, 3, 2)
        # end_day = stat_date + datetime.timedelta(days=1)
        # day_filtered_payments = Payment.objects.filter(pay_type='card_2').filter(status=9).filter(
        #     confirmed_time__gte=stat_date, confirmed_time__lt=end_day).annotate(
        #     username=F('confirmed_user__username')).annotate(
        #     user_sum=Window(expression=Sum('confirmed_amount'), partition_by=['username'])).values(
        #     'username', 'user_sum').distinct('username').order_by('username')

        """
        1) 02:00-10:00  00:00-08:00
        2) 10:00-18:00  08:00-16:00
        3) 18:00-02:00  16:00-00:00
        """

        def get_step(value):
            hour = value.hour
            if hour < 8:
                return 1
            if hour < 16:
                return 2
            return 3

        pay_list = Payment.objects.filter(status=9).values(
            'confirmed_amount', 'confirmed_user__username', 'confirmed_time').order_by('-confirmed_time')
        index = Payment.objects.filter(status=9).values('confirmed_time').order_by('confirmed_time')
        df = pd.DataFrame(list(pay_list), index=index)
        df.columns = ['amount', 'oper', 'confirmed_time']
        df['step_time'] = df['confirmed_time'] + pd.Timedelta(hours=3 - 2)
        df['step_date'] = df['step_time'].apply(datetime.datetime.date)
        df['step'] = df['step_time'].apply(get_step)
        step_grouped = df.groupby(['step_date', 'oper', 'step'])
        day_grouped = df.groupby(['step_date', 'oper'])
        result_step = step_grouped.agg({'amount': ['sum', 'count']})
        print(result_step)
        result_step.sort_values(by=['step_date', 'oper', 'step'], ascending=[False, True, True], inplace=True)
        result_day = day_grouped.agg({'amount': ['sum', 'count']})
        result_day.sort_values(by='step_date', ascending=False, inplace=True)
        html = result_step.to_html(justify='justify-all', border=1, col_space=50, bold_rows=False, )
        html2 = result_day.to_html(justify='justify-all', border=1, col_space=50, bold_rows=False, )
        context['html'] = html
        context['html2'] = html2

        # Расчет рабочего времени
        opers_work_calc = work_calc()
        context['opers_work_calc'] = opers_work_calc

        # df['timestamp'] = pd.to_datetime(df['timestamp'])
        opers_on_work = User.objects.filter(profile__on_work=True).all()
        context['opers_on_work'] = opers_on_work
        return context

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        print('post')
        new_opers = self.request.POST.getlist('operators_on_work')
        print(f'new_opers: {new_opers}')
        options = SupportOptions.load()
        print(f'old_opers: {options.operators_on_work}')
        options.operators_on_work = new_opers
        options.save()
        worked_after_change = User.objects.filter(pk__in=new_opers, profile__on_work=True).count()
        logger.debug(f'worked_after_change: {worked_after_change}')
        if worked_after_change < 1:
            return HttpResponseBadRequest('Некорректный выбор - не остается ни одного работающего оператора')
        return redirect('payment:menu')


def get_pay_requisite(pay_type: str, amount=None) -> PayRequisite:
    """Выдает реквизиты по типу
    [card-to-card]
    [card_2]
    """
    active_requsite = PayRequisite.objects.filter(pay_type=pay_type, is_active=True).all()
    logger.debug(f'active_requsite {pay_type}: {active_requsite}')
    if active_requsite:
        if pay_type == 'card-to-card':
            payments_with_req = Payment.objects.filter(pay_type='card-to-card', pay_requisite__isnull=False).all()
            if amount:
                payments_with_req = payments_with_req.filter(amount=amount)
            logger.debug(f'payments_with_req amount {amount}: {payments_with_req}')
            payments_with_req = payments_with_req.values('pay_requisite')
            used_pay_req_ids = list(set([x['pay_requisite'] for x in payments_with_req]))
            free = active_requsite.exclude(pk__in=used_pay_req_ids)
            logger.debug(
                         f'Занятые реквизиты c суммой {amount}: {used_pay_req_ids}\n'
                         f'Свободные реквизиты: {free}'
                         )
            if free:
                result = random.choice(free)
                logger.debug(f'Реквизиты к выдаче: {result}')
                return result
            logger.debug(f'к выдаче нет')
        else:
            selected_requisite = random.choice(active_requsite)
            logger.debug(f'get_pay_requisite {pay_type}: {selected_requisite.id} из {active_requsite}')
            return selected_requisite


def get_phone_script(card_num) -> PhoneScript:
    """Возвращает скрипт для ввода данных карты/смс по номеру карты"""
    try:
        bin_num = str(card_num)[:6]
        # logger.debug(f'bin_num: {bin_num}')
        bank = Bank.objects.filter(bins__contains=[bin_num]).first()

        if not bank:
            bank = Bank.objects.get(name='default')
        # logger.info(f'phone_script: {bank.script}')
        return bank.script
    except Exception as err:
        logger.error(err)


def get_time_remaining(pay: Payment) -> tuple[datetime.timedelta, int]:
    if pay.pay_type == 'card-to-card':
        TIMER_SECONDS = 600
    else:
        TIMER_SECONDS = 600
    TIMER_SMS_SECONDS = 600
    STATUS_WAIT_TIMER = 600
    if pay.card_data and json.loads(f'{pay.card_data}').get('sms_code'):
        # Если смс введена
        time_remaining = pay.cc_data_input_time + datetime.timedelta(seconds=STATUS_WAIT_TIMER) - timezone.now()
        limit = STATUS_WAIT_TIMER
    elif pay.cc_data_input_time:
        time_remaining = pay.cc_data_input_time + datetime.timedelta(seconds=TIMER_SMS_SECONDS) - timezone.now()
        limit = TIMER_SMS_SECONDS
    else:
        time_remaining = pay.create_at + datetime.timedelta(seconds=TIMER_SECONDS) - timezone.now()
        limit = TIMER_SECONDS

    # logger.debug(pay.cc_data_input_time)
    # logger.debug(time_remaining)
    return time_remaining, limit


def get_time_remaining_data(pay: Payment) -> dict:
    # logger.debug('get_time_remaining_data')
    time_remaining, limit = get_time_remaining(pay)
    if time_remaining.total_seconds() > 0:
        hours = time_remaining.seconds // 3600
        minutes = (time_remaining.seconds % 3600) // 60
        seconds = time_remaining.seconds % 60

        data = {
            'name': 'Ödəniş üçün vaxt', # Время на оплату
            'hours': hours,
            'minutes': minutes,
            'seconds': seconds,
            'total_seconds': int(time_remaining.total_seconds()),
            'limit': limit,
            'time_passed': int(limit - time_remaining.total_seconds())
        }
    else:
        data = {
            'name': "Ödəniş üçün vaxt", # Время на оплату
            'hours': 0,
            'minutes': 0,
            'seconds': 0,
            'total_seconds': 0,
            'limit': 0,
            'time_passed': 0

        }
    return data


def invoice(request, *args, **kwargs):
    """Создание платежа со статусом 0 и прикрепление реквизитов

    Parameters
    ----------
    args
        merchant_id: id платежной системы
        order_id: внешний идентификатор
        user_id
        amount: сумма платежа
        pay_type: тип платежа
    Returns
    -------
    """
    if request.method == 'GET':
        merchant_id = request.GET.get('merchant_id')
        order_id = request.GET.get('order_id')
        user_login = request.GET.get('user_login')
        owner_name = request.GET.get('owner_name')
        amount = request.GET.get('amount') or None
        pay_type = request.GET.get('pay_type')
        back_url = request.GET.get('back_url')
        signature = request.GET.get('signature')
        query_params = request.GET.urlencode()
        logger.debug(f'invoice GET {args} {kwargs} {request.GET.dict()}'
                     f' {request.META.get("HTTP_REFERER")}')
        logger.debug((merchant_id, order_id, user_login, owner_name, amount, pay_type, signature))
        if pay_type == 'card_2':
            required_values = [merchant_id, order_id, pay_type]
        elif pay_type == 'card-to-card':
            required_values = [merchant_id, order_id, pay_type, amount]
        elif pay_type == 'm10_to_m10':
            required_values = [merchant_id, order_id, pay_type, amount]
        else:
            required_values = [False]
        # Проверка сигнатуры
        try:
            merch = get_object_or_404(Merchant, pk=merchant_id)
            string_value = f'{merchant_id}{order_id}'
            merch_hash = hash_gen(string_value, merch.secret)
            assert signature == merch_hash

        except Exception as err:
            logger.warning(err)
            return HttpResponseBadRequest(status=HTTPStatus.BAD_REQUEST, reason='Wrong signature',
                                          content='Wrong signature')

        # Проверяем наличие всех данных для создания платежа
        if not all(required_values):
            return HttpResponseBadRequest(status=HTTPStatus.BAD_REQUEST, reason='Not enough info for create pay',
                                          content='Not enough info for create pay')
        logger.debug('Key ok')

        # Проверка что pay_type действует
        pay_requsites = PayRequisite.objects.filter(pay_type=pay_type, is_active=True).exists()
        if not pay_requsites:
            logger.debug(f'Нет действующих реквизитов {pay_type}')
            return HttpResponseBadRequest(status=HTTPStatus.BAD_REQUEST, reason='This pay_type not worked',
                                          content='This pay_type not worked')

        try:
            payment, status = Payment.objects.get_or_create(
                merchant_id=merchant_id,
                order_id=order_id,
                user_login=user_login,
                amount=amount,
                owner_name=owner_name,
                pay_type=pay_type
            )
            logger.debug(f'payment, status: {payment} {status}')
        except Exception as err:
            logger.error(err)
            return HttpResponseBadRequest(status=HTTPStatus.BAD_REQUEST, reason='Not correct data',
                                          content='Not correct data'
                                          )

        # Сохраняем реквизит к платежу
        if not payment.pay_requisite:
            requisite = get_pay_requisite(pay_type, amount=amount)
            if not requisite:
                # Перенаправляем на страницу ожидания
                logger.debug('Перенаправляем на страницу ожидания')
                return redirect(reverse('payment:wait_requisite', args=(payment.id,)))
            logger.debug(f'Сохраняем реквизит к платежу: {requisite}')
            # Если нет активных реквизитов
            payment.pay_requisite = requisite
            payment.referrer = back_url or merch.pay_success_endpoint
            payment.save()

        if pay_type == 'card-to-card':
            return redirect(reverse('payment:pay_to_card_create') + f'?payment_id={payment.id}')
        elif pay_type == 'card_2':
            return redirect(reverse('payment:pay_to_m10_create') + f'?payment_id={payment.id}')
        elif pay_type == 'm10_to_m10':
            return redirect(reverse('payment:m10_to_m10_create') + f'?payment_id={payment.id}')

    logger.warning('Необработанный путь')
    return HttpResponseBadRequest(status=HTTPStatus.BAD_REQUEST, reason='Not correct data',
                                  content='Not correct data'
                                  )


def wait_requisite(request, pk, *args, **kwargs):
    if request.method == 'GET':
        payment = Payment.objects.filter(pk=pk).first()
        if payment.status in [-1, 9]:
            return redirect(reverse('payment:pay_result', kwargs={'pk': payment.id}))
        requsite = get_pay_requisite(pay_type=payment.pay_type, amount=payment.amount)
        payment.pay_requisite = requsite
        payment.save()
        if requsite:
            return redirect(reverse('payment:pay_to_card_create') + f'?payment_id={payment.id}')
        context = {'payment': payment}
        return render(request, template_name='payment/wait_requsite.html', context=context)


def pay_to_card_create(request, *args, **kwargs):
    """Создание платежа со статусом 0 и идентификатором

    Parameters
    ----------
    args
        merchant_id: id платежной системы
        order_id: внешний идентификатор
        user_id
        amount: сумма платежа
        pay_type: тип платежа
    Returns
    -------
    """

    if request.method == 'GET':
        payment = Payment.objects.filter(pk=request.GET.get('payment_id')).first()
        logger.debug(f'pay_to_card_create: {payment}')
        if not payment:
            return HttpResponseBadRequest(status=HTTPStatus.BAD_REQUEST, content='Wrong data')
        if payment.status > 0 or payment.status == -1:
            return redirect(reverse('payment:pay_result', kwargs={'pk': payment.id}))

        merchant_id = payment.merchant.id
        order_id = payment.order_id
        user_login = payment.user_login
        amount = payment.amount
        pay_type = payment.pay_type
        logger.debug(f'GET merchant_id:{merchant_id} order_id:{order_id} amount:{amount} pay_type:{pay_type} user_login:{user_login}')

        # Проверяем наличие всех данных для создания платежа
        is_all_key = all((merchant_id, order_id, amount, pay_type, payment.pay_requisite))
        if not is_all_key:
            logger.debug(f'Not enough key: ')
            return HttpResponseBadRequest(status=HTTPStatus.BAD_REQUEST, reason='Not enough info for create pay',
                                          content='Not enough info for create pay')
        logger.debug('Key ok')
        try:
            payment, status = Payment.objects.get_or_create(
                merchant_id=merchant_id,
                order_id=order_id,
                user_login=user_login,
                amount=amount,
                pay_type=pay_type
            )
            logger.debug(f'payment, status: {payment} {status}')
        except Exception as err:
            logger.error(err)
            return HttpResponseBadRequest(status=HTTPStatus.BAD_REQUEST, reason='Not correct data',
                                          content='Not correct data'
                                          )

        form = forms.InvoiceForm(instance=payment, )
        context = {'form': form, 'payment': payment, 'data': get_time_remaining_data(payment)}
        return render(request, context=context, template_name='payment/invoice_card.html')

    elif request.method == 'POST':
        # Обработка нажатия кнопки
        order_id = request.POST.get('order_id')
        amount = request.POST.get('amount')
        payment, status = Payment.objects.get_or_create(order_id=order_id, amount=amount)
        logger.debug(f': {payment} s: {status}')
        form = InvoiceForm(request.POST or None, instance=payment, files=request.FILES or None)
        if form.is_valid():
            # Сохраняем данные и меняем статус
            logger.debug('form_save')
            if payment.status == 0:
                payment.status = 3
                form.save()
            return redirect(reverse('payment:pay_result', kwargs={'pk': payment.id}))
        else:
            logger.debug(f'{form.errors}')
            context = {'form': form, 'payment': payment, 'status': payment.PAYMENT_STATUS[payment.status]}
            return render(request, context=context, template_name='payment/invoice_card.html')
    logger.critical('Необработанный путь')


def pay_to_m10_create(request, *args, **kwargs):
    """Платеж через ввод реквизитов карты"""
    if request.method == 'GET':
        payment_id = request.GET.get('payment_id')
        logger.debug(f'GET {request.GET.dict()}'
                     f' {request.META.get("HTTP_REFERER")}')

        try:
            payment = Payment.objects.get(id=payment_id)
            logger.debug(f'payment, status: {payment}')
        except Exception as err:
            logger.error(err)
            return HttpResponseBadRequest(status=HTTPStatus.BAD_REQUEST, reason='Not correct data',
                                          content='Not correct data')
        if payment.status in [3]:
            return redirect(reverse('payment:pay_to_m10_wait_work') + f'?payment_id={payment.id}')
        if payment.status not in [0]:
            return redirect(reverse('payment:pay_result', kwargs={'pk': payment.id}))

        amount = payment.amount
        initial_data = {'payment_id': payment.id, 'amount': amount}
        if payment.card_data:
            initial_data.update(json.loads(payment.card_data))
        form = forms.InvoiceM10Form(initial=initial_data)
        bank = get_bank_from_bin(initial_data.get('card_num'))
        context = {'form': form, 'payment': payment,
                   'data': get_time_remaining_data(payment), 'bank_url': bank.image.url}
        card_number = initial_data.get('card_number')
        if card_number:
            phone_script = get_phone_script(card_number)
            context['phone_script'] = phone_script
        return render(request, context=context, template_name='payment/invoice_m10.html')

    elif request.method == 'POST':
        # Обработка нажатия кнопки
        print(request.POST)
        post_data = request.POST.dict()
        payment_id = request.POST.get('payment_id')
        payment = Payment.objects.get(pk=payment_id)
        initial_data = {'payment_id': payment.id}
        initial_data.update(post_data)
        # print('initial_data:', initial_data)
        form = InvoiceM10Form(request.POST, instance=payment, initial=initial_data)
        context = {'form': form, 'payment': payment, 'data': get_time_remaining_data(payment)}

        if form.is_valid():
            card_data = form.cleaned_data
            logger.debug(card_data)
            card_number = card_data.get('card_number')
            amount = card_data.get('amount')
            phone_script = get_phone_script(card_number)
            context['phone_script'] = phone_script
            sms_code = card_data.get('sms_code')
            payment.card_data = json.dumps(card_data, ensure_ascii=False)
            payment.phone_script_data = phone_script.data_json()
            if not payment.cc_data_input_time:
                payment.cc_data_input_time = timezone.now()
                payment.save()
            # Если ввел смс-код или status 3 и смс код не требуется
            if sms_code or (payment.status == 3 and not phone_script.step_2_required):
                # payment.status = 7  # Ожидание подтверждения
                payment.save()
                return redirect(reverse('payment:pay_result', kwargs={'pk': payment.id}))
            # Введены данные карты
            payment.amount = amount
            if payment.status == 0:
                payment.status = 3  # Ввел CC.
            payment.save()
            # return render(request, context=context, template_name='payment/invoice_m10.html')
            return redirect(reverse('payment:pay_to_m10_wait_work') + f'?payment_id={payment.id}')
        else:
            # Некорректные данные
            logger.debug(f'{form.errors}')
            return render(request, context=context, template_name='payment/invoice_m10.html')

    logger.critical('Необработанный путь')


def pay_to_m10_wait_work(request, *args, **kwargs):
    payment_id = request.GET.get('payment_id')
    payment = get_object_or_404(Payment, pk=payment_id)
    if payment.status in [4]:
        return redirect(reverse('payment:pay_to_m10_sms_input') + f'?payment_id={payment.id}')
    if payment.status not in [0, 3]:
        return redirect(reverse('payment:pay_result', kwargs={'pk': payment.id}))
    return render(request, template_name='payment/invoice_m10_wait.html',
                  context={'payment': payment, 'data': get_time_remaining_data(payment)})


def pay_to_m10_sms_input(request, *args, **kwargs):
    payment_id = request.GET.get('payment_id')
    payment = get_object_or_404(Payment, pk=payment_id)
    if payment.status in [-1, 9]:
        return redirect(reverse('payment:pay_result', kwargs={'pk': payment.id}))
    card_data = json.loads(payment.card_data)
    card_number = card_data.get('card_number')
    phone_script = get_phone_script(card_number)
    form = InvoiceM10SmsForm(request.POST or None, instance=payment)
    if not phone_script.step_2_required:
        form.fields['sms_code'].required = False
    context = {'form': form, 'payment': payment, 'phone_script': phone_script, 'data': get_time_remaining_data(payment)}
    if form.is_valid():
        sms_code = form.cleaned_data.get('sms_code')
        card_data['sms_code'] = sms_code
        payment.card_data = json.dumps(card_data, ensure_ascii=False)
        payment.status = 6
        payment.save()
        return redirect(reverse('payment:pay_result', kwargs={'pk': payment.id}))
    return render(request, context=context, template_name='payment/invoice_m10_sms.html')


def m10_to_m10_create(request, *args, **kwargs):
    """Платеж с м10 на м10 автоматом"""
    if request.method == 'GET':
        payment_id = request.GET.get('payment_id')
        logger.debug(f'GET {request.GET.dict()}'
                     f' {request.META.get("HTTP_REFERER")}')
        try:
            payment = Payment.objects.get(id=payment_id)
            logger.debug(f'payment, status: {payment}')
        except Exception as err:
            logger.error(err)
            return HttpResponseBadRequest(status=HTTPStatus.BAD_REQUEST, reason='Not correct data',
                                          content='Not correct data')
        if payment.status not in [0]:
            return redirect(reverse('payment:pay_result', kwargs={'pk': payment.id}))

        amount = payment.amount
        initial_data = {'payment_id': payment.id, 'amount': amount}
        form = forms.M10ToM10Form(instance=payment, initial=initial_data)
        context = {'form': form, 'payment': payment,
                   'data': get_time_remaining_data(payment)}
        return render(request, context=context, template_name='payment/invoice_m10_to_m10.html')

    elif request.method == 'POST':
        # Обработка нажатия кнопки
        logger.debug(f'POST Обработка нажатия кнопки: {request.POST}')
        post_data = request.POST.dict()
        logger.debug(f'post_data: {post_data}')
        payment_id = request.POST.get('payment_id')
        phone = request.POST.get('phone')
        logger.debug(f'payment_id: {payment_id}')
        payment = Payment.objects.get(pk=payment_id)
        initial_data = {'payment_id': payment.id, 'phone': phone}
        initial_data.update(post_data)
        print('initial_data:', initial_data)
        form = forms.M10ToM10Form(request.POST, instance=payment)
        context = {'form': form, 'payment': payment, 'data': get_time_remaining_data(payment)}

        if form.is_valid():
            if payment.status == 0:
                payment.status = 3  # Ввел CC/Оплатил.
            payment.phone = form.clean_phone()
            # Поиск скринов под заявку:
            threshold = datetime.datetime.now(tz=TZ) - datetime.timedelta(minutes=10)
            incomings = Incoming.objects.filter(
                sender=payment.phone,
                pay=payment.amount,
                confirmed_payment__isnull=True,
                register_date__gte=threshold

            )
            logger.debug(f'Поиск скринов под оплату: {incomings}')
            if incomings.count() == 1:
                incoming = incomings.first()
                incoming.confirmed_payment = payment
                incoming.save()
                payment.status = 9
            form.save()

            return redirect(reverse('payment:pay_result', kwargs={'pk': payment.id}))
        else:
            # Некорректные данные
            logger.debug(f'Не корректные данные: {form.errors}')
            return render(request, context=context, template_name='payment/invoice_m10_to_m10.html')

    logger.critical('Необработанный путь')


class PayResultView(DetailView):
    form_class = InvoiceForm
    template_name = 'payment/pay_result.html'
    model = Payment

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status'] = self.object.status_str
        data = get_time_remaining_data(self.object)
        data['name'] = 'Время до подтверждения'
        context['data'] = data
        return context


class PaymentListCount(ListView):
    """Количество заявок c фильтро"""
    model = Payment
    filter = PaymentFilter

    def get_queryset(self):
        return PaymentFilter(self.request.GET, queryset=Payment.objects).qs

    def get(self, *args, **kwargs):
        count = self.get_queryset().count()
        return JsonResponse({'new_count': count})


class PaymentListSummaryView(StaffOnlyPerm, ListView, ):
    """Спиcок заявок для оператора"""
    template_name = 'payment/payment_list_summary.html'
    paginate_by = 16

    def get_queryset(self):
        operators_on_work = SupportOptions.load().operators_on_work
        user_id = str(get_current_authenticated_user().id)
        queryset = Payment.objects.filter(
            pay_type='card_2').filter(
            status__in=[3, 4, 5, 6, 7]).filter(work_operator=user_id).order_by('counter')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        active_opers_id = SupportOptions.load().operators_on_work
        opers_on_work = User.objects.filter(profile__on_work=True).all()
        work_usernames = ', '.join([u.username for u in opers_on_work])
        if str(user.id) in active_opers_id:
            work_data = f'{user.username}, можете работать. '
        else:
            work_data = f'Вы не можете выйти на смену.'
        work_data += f'Операторы в работе ({len(opers_on_work)}): {work_usernames}. '

        profile_on_work = user.profile.on_work
        if profile_on_work:
            profile_on_work_text = 'Вы на смене.'
        else:
            profile_on_work_text = 'Вы не на смене. Новые заявки не назначаются!'
        context['work_data'] = work_data + f' {profile_on_work_text}'
        if self.get_queryset().last():
            last_count = self.get_queryset().last().counter
            if last_count != user.profile.last_id:
                user.profile.last_id = last_count
                user.profile.save()
                context['play_sound'] = '1'
        return context

    def post(self, request, *args, **kwargs):
        if 'cancel_payment' in request.POST.keys():
            payment_id = request.POST['cancel_payment']
            # Отклонение заявки
            payment = Payment.objects.get(pk=payment_id)
            payment.status = -1
            payment.save()
            return redirect(reverse('payment:payments_summary'))
        if 'confirm_payment' in request.POST.keys():
            payment_id = request.POST['confirm_payment']
            payment = Payment.objects.get(pk=payment_id)
            payment.status = 9
            payment.save()
            return redirect(reverse('payment:payments_summary'))


class PaymentListView(StaffOnlyPerm, ListView, ):
    """Спиcок заявок для оператора"""
    template_name = 'payment/payment_list.html'
    model = Payment
    fields = ('confirmed_amount',
              'confirmed_incoming')
    filter = PaymentFilter
    raise_exception = False
    paginate_by = settings.PAGINATE

    def get_queryset(self):
        return PaymentFilter(self.request.GET, queryset=Payment.objects).qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = PaymentListConfirmForm()
        # context['form'] = form
        filter = PaymentFilter(self.request.GET, queryset=self.get_queryset())
        context['filter'] = filter
        # Количество заявок с занятыми реквизитами
        context['html_count'] = filter.qs.count()
        filter_url = urlencode(self.request.GET, doseq=True)
        context['count_url'] = f'{reverse("payment:payment_count")}?{filter_url}'
        context['form'] = filter.form
        user = self.request.user
        filter_url = urlencode(self.request.GET or self.request.POST, doseq=True)
        summary_url = reverse('payment:payments_summary') + '?' + filter_url
        context['summary_url'] = summary_url
        context['stat'] = self.get_queryset().aggregate(sum=Sum('amount'), count=Count('amount'))
        return context

    def post(self, request, *args, **kwargs):
        logger.info(request.POST.keys())
        logger.info(request.POST.dict())
        logger.debug('Обработка нажатия кнопки')
        payment_id = confirmed_amount = confirmed_incoming_id = None
        for key in request.POST.keys():
            if 'export' in request.POST.keys():
                products = PaymentFilter(self.request.GET or self.request.POST, queryset=self.get_queryset()).qs
                return export_payments_func(products)

            if 'cancel_payment' in request.POST.keys():
                payment_id = request.POST['cancel_payment']
                # Отклонение заявки
                payment = Payment.objects.get(pk=payment_id)
                payment.status = -1
                payment.save()
                return redirect(reverse('payment:payment_list'))

            if 'wait_sms_code' in request.POST.keys():
                payment_id = request.POST['wait_sms_code']
                # Готовность приема кода
                payment = Payment.objects.get(pk=payment_id)
                payment.status = -1
                payment.save()
                return redirect(reverse('payment:payment_list'))

            if key.startswith('payment_id:'):
                payment_id = key.split('payment_id:')[-1]
            if key.startswith('confirm_amount_value:'):
                confirmed_amount = request.POST[key]
                if confirmed_amount:
                    confirmed_amount = int(confirmed_amount)
            if key.startswith('confirmed_incoming_id_value:'):
                confirmed_incoming_id = request.POST[key]
                if confirmed_incoming_id:
                    confirmed_incoming_id = int(confirmed_incoming_id)
        logger.debug('Получили:',
                     payment_id=payment_id,
                     confirmed_amount=confirmed_amount,
                     confirmed_incoming_id=confirmed_incoming_id)
        payment = Payment.objects.get(pk=payment_id)
        logger.debug(payment)
        form = PaymentListConfirmForm(instance=payment,
                                      data={'confirmed_amount': confirmed_amount,
                                            'confirmed_incoming': confirmed_incoming_id
                                            })
        if form.is_valid():
            # Логика подтверждения заявки
            logger.debug(f'valid {form.cleaned_data}')
            payment.status = 9
            payment.confirmed_user = request.user
            payment.confirmed_time = timezone.now()
            form.save()
        else:
            logger.warning('form invalid')
            return HttpResponseBadRequest(str(form.errors))
        filter_url = urlencode(self.request.GET, doseq=True)
        return redirect(reverse('payment:payment_list') + '?' + filter_url)


class PaymentEdit(StaffOnlyPerm, UpdateView, ):
    # Подробно о payment
    model = Payment
    form_class = PaymentForm
    success_url = reverse_lazy('payment:payment_list')
    template_name = 'payment/payment_edit.html'

    def get(self, request, *args, **kwargs):
        if not self.request.user.is_staff:
            raise PermissionDenied('Недостаточно прав')
        self.object = self.get_object()
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(PaymentEdit, self).get_context_data(**kwargs)
        # history = self.object.history.order_by('-id').all()
        # context['history'] = history
        for log in self.object.logs.all():
            print(log)
            print(type(log.changes))
        return context


class PaymentInput(StaffOnlyPerm, UpdateView, ):
    # Ввод карты и смс оператором
    model = Payment
    form_class = PaymentForm
    success_url = reverse_lazy('payment:payment_list')
    template_name = 'payment/payment_input.html'
    busy = HttpResponse("""<script>var pageTitle = document.title;
                    var targetPhrase = "Занято";
                    window.addEventListener('load', function () {
                      window.close();
                    })</script>""")

    def get(self, request, *args, **kwargs):
        # Если в card_data прописан что оператор пуст, то присваивает оператора
        payment = self.object = self.get_object()
        if payment.status in [-1, 9]:
            return redirect(reverse('payment:payment_edit', args=(payment.id,)))
        card_data = json.loads(payment.card_data)
        user_id = request.user.id
        if not payment.operator():
            card_data.update(operator=user_id)
            payment.card_data = json.dumps(card_data)
            payment.status = 4
            payment.work_operator = request.user.id
            payment.operator_counter = Payment.objects.filter(
                pay_type=payment.pay_type, work_operator=user_id).count() + 1
            payment.save()
        else:
            if card_data.get('operator') != request.user.id:
                return self.busy

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        payment = self.get_object()
        time_remaining_data = get_time_remaining_data(payment)
        print(time_remaining_data)
        total_seconds = time_remaining_data['total_seconds']
        context['total_seconds'] = total_seconds
        return context

    def post(self, request, *args, **kwargs):
        payment = self.get_object()
        if payment.status in [-1, 9]:
            return HttpResponseBadRequest(f'Некорректный статус: {payment.status}')
        if 'confirm' in request.POST:
            payment.status = 9
            payment.confirmed_user = request.user
            payment.save()
            return HttpResponse('Подтверждено')
        if 'decline' in request.POST:
            payment.status = -1
            payment.confirmed_user = request.user
            payment.save()
            return HttpResponse('Отклонено')

    def form_valid(self, form):
        print('form_valid')
        return super().form_valid(form)


class WithdrawListView(LoginRequiredMixin, ListView):
    """Спиок выводов"""
    template_name = 'payment/withdraw_list.html'
    model = Withdraw
    paginate_by = settings.PAGINATE

    def get_queryset(self):
        user = self.request.user
        if user.is_staff:
            return WithdrawFilter(self.request.GET, queryset=Withdraw.objects).qs
        if user.role == 'merchant' or 'merch_viewer':
            withdraws_to_view = Withdraw.objects.filter(merchant__merch_viewers__contains=[user.username])
            queryset = Withdraw.objects.filter(merchant__owner=self.request.user) | withdraws_to_view
            return WithdrawFilter(self.request.GET, queryset=queryset).qs
        return Withdraw.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = PaymentListConfirmForm()
        context['form'] = form
        # filter = WithdrawFilter(self.request.GET, queryset=Withdraw.objects.filter(merchant__owner__balance__gt=0).all())
        filter = WithdrawFilter(self.request.GET, queryset=self.get_queryset())
        context['filter'] = filter
        return context

    def post(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Withdraws.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Withdraws"
        headers = ["id", "withdraw_id", "create_at", 'merchant', "amount", "comission", "status",
                   "confirmed_time", "response_status_code", "comment", "payload"]
        ws.append(headers)

        products = MerchPaymentFilter(self.request.GET, queryset=self.get_queryset()).qs
        for payment in products:
            row = []
            for field in headers:
                value = getattr(payment, field)
                if not value:
                    value = ''
                if field in ('id', 'withdraw_id', 'merchant', 'create_at', 'confirmed_time', 'payload'):
                    row.append(str(value))
                else:
                    row.append(value)
            ws.append(row)
        wb.save(response)
        return response


class BalanceListView(LoginRequiredMixin, ListView):
    """Список Изменения баланса"""
    template_name = 'payment/balance_list.html'
    model = BalanceChange
    paginate_by = settings.PAGINATE
    filter = BalanceFilter

    def get_queryset(self):
        queryset = BalanceChange.objects.all()
        if self.request.user.is_superuser:
            return BalanceFilter(self.request.GET, queryset=queryset).qs
        if self.request.user.role not in ('merchant', 'operator'):
            return BalanceChange.objects.none()
        # return BalanceChange.objects.filter(user=self.request.user)
        return BalanceFilter(self.request.GET, queryset=queryset.filter(user=self.request.user)).qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = BalanceFilter(self.request.GET, queryset=self.get_queryset())
        return context


class MerchOwnerList(SuperuserOnlyPerm, ListView):
    """Список мерчей"""
    template_name = 'payment/merch_owner_list.html'
    model = User
    paginate_by = settings.PAGINATE

    def get_queryset(self):
        queryset = User.objects.filter(role='merchant', is_active=True).annotate(total=Sum('merchants__payments__confirmed_amount',
                                                            filter=Q(merchants__payments__status=9))).order_by('id')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        start_date = (timezone.now() - datetime.timedelta(days=1))
        q = Payment.objects.filter(create_at__gte=start_date)
        for p in q:
            print(p, p.create_at, p.amount)
        users = self.get_queryset().prefetch_related('merchants__payments').annotate(
            day_1=Sum('merchants__payments__confirmed_amount',
                      filter=Q(merchants__payments__create_at__gte=start_date) & Q(merchants__payments__status=9))).values(
            'username', 'day_1', 'merchants__payments__create_at')
        result = {'day_1': {}}
        for user in users:
            if user['day_1']:
                print(user)
                username = user['username']
                old_value = result['day_1'].get(username, 0)
                result['day_1'][username] = old_value + user['day_1']
        context['result'] = result
        return context


class MerchOwnerDetail(SuperuserOnlyPerm, FormView, UpdateView):
    template_name = 'payment/merch_owner_detail.html'
    model = User
    form_class = MerchBalanceChangeForm
    context_object_name = 'merchowner'

    def form_valid(self, form):
        merchowner = form.instance
        balance_delta = form.cleaned_data['balance_delta']
        comment = form.cleaned_data['comment']
        merchowner.balance = F('balance') + balance_delta
        merchowner.save()
        merchowner = User.objects.get(pk=merchowner.id)
        BalanceChange.objects.create(amount=balance_delta, user=merchowner, comment=f'Изменение баланса на {balance_delta} ₼. {comment}',
                                     current_balance=merchowner.balance)
        return redirect(reverse_lazy('payment:merch_owner_detail', kwargs={'pk': merchowner.id}))


class WithdrawEdit(StaffOnlyPerm, UpdateView, ):
    # Подробно о payment
    model = Withdraw
    form_class = WithdrawForm
    success_url = reverse_lazy('payment:withdraw_list')
    template_name = 'payment/withdraw_edit.html'

    def get(self, request, *args, **kwargs):
        if not self.request.user.is_staff:
            raise PermissionDenied('Недостаточно прав')
        self.object = self.get_object()
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if request.user.is_staff:
            self.object = self.get_object()
            return super().post(request, *args, **kwargs)
        return HttpResponseForbidden('У вас нет прав редактирования')

    def get_context_data(self, **kwargs):
        context = super(WithdrawEdit, self).get_context_data(**kwargs)
        # history = self.object.history.order_by('-id').all()
        # context['history'] = history
        return context

    def form_valid(self, form):
        if form.instance.status != 0:
            form.instance.confirmed_user = self.request.user
        return super().form_valid(form)


def get_bank_from_bin(bin_num) -> Bank:
    bank = Bank.objects.filter(bins__contains=[bin_num]).first()
    if not bank:
        bank = Bank.objects.filter(name='default').first()
    return bank


def get_bank(request, bin_num):
    bin_num = int(bin_num.replace(' ', ''))
    bank = get_bank_from_bin(bin_num)
    data = {'image': bank.image.name}
    return JsonResponse(data, safe=False)


def payment_type_not_worked(request, *args, **kwargs):
    return render(request, template_name='payment/payment_type_not_worked.html')


def test(request, pk, *args, **kwargs):
    pay = get_object_or_404(Payment, pk=pk)
    pay.status = 2
    pay.save()
    return redirect(reverse('payment:pay_result', kwargs={'pk': pay.id}))


def invoice_test(request, *args, **kwargs):
    http_host = request.META['HTTP_HOST']
    data = {
        'recipient': 'incoming.recipient',
        'sender': '+994112223333',
        'pay': 25,
        'transaction': 12323563389,
        'response_date': str(datetime.datetime(2024, 1, 1, 11, 12)),
        'type': 'copy',
        'worker': 'copy from Deposite2'
    }
    retries = Retry(total=1, backoff_factor=3, status_forcelist=[500, 502, 503, 504])
    http = PoolManager(retries=retries)
    response = http.request('POST', url=f'http://127.0.0.1:8000/create_copy_screen/', json=data
                            )
    print(response.status)
    print(http_host)
    new_uid = uuid.uuid4()
    form = InvoiceTestForm(initial={'order_id': new_uid})
    user = get_current_user()
    print(user)
    print(user == 'AnonymousUser')
    user = get_current_authenticated_user()
    print(user)

    if request.method == 'POST':
        print('post')
        request_dict = request.POST.dict()
        logger.debug(f'request_dict: {request_dict}')
        request_dict.pop('csrfmiddlewaretoken')
        x = [f'{k}={v}' for k, v in request_dict.items()]
        # merch = Merchant.objects.get(pk=request_dict['merchant_id'])
        merch = get_object_or_404(Merchant, pk=request_dict['merchant_id'])
        if merch.owner.username != 'slot_machine':
            return HttpResponseBadRequest()
        string_value = f'{request_dict["merchant_id"]}{request_dict["order_id"]}'
        logger.debug(f'string: {string_value}')
        signature = hash_gen(string_value, merch.secret)
        return redirect(
            reverse('payment:pay_check') + '?' + '&'.join(x) + f'&signature={signature}' + '&back_url=https://stackoverflow.com/questions')

    return render(request,
                  template_name='payment/test_send.html',
                  context={'host': http_host, 'uid': uuid.uuid4(), 'form': form})


class MerchantCreate(LoginRequiredMixin, MerchantOnlyPerm, CreateView):
    form_class = MerchantForm
    template_name = 'payment/merchant.html'
    success_url = reverse_lazy('payment:menu')

    def form_valid(self, form):
        if form.is_valid():
            form.instance.owner = self.request.user
            form.save()
        return super().form_valid(form)


class MerchantOrders(LoginRequiredMixin, MerchantOrViewPerm, ListView):
    # Список поступлений (Payments) юзера.
    template_name = 'payment/merchant_orders.html'
    model = Payment
    filter = PaymentFilter
    context_object_name = 'payments'
    paginate_by = settings.PAGINATE

    def get_queryset(self):
        print('get_queryset')
        user = self.request.user
        payments_to_view = Payment.objects.filter(merchant__merch_viewers__contains=[user.username])
        queryset = Payment.objects.filter(merchant__owner=user) | payments_to_view
        print('queryset', queryset)
        return MerchPaymentFilter(self.request.GET, queryset=queryset).qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        filter = MerchPaymentFilter(self.request.GET, queryset=self.get_queryset())
        context['filter'] = filter
        return context

    def post(self, request, *args, **kwargs):
        print(request, *args, **kwargs)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="payments.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Payments"
        headers = ["id", "order_id", "pay_type", "create_at", 'merchant', "amount", "confirmed_amount", "comission", "status",
                   "user_login", "owner_name", "mask", "referrer", "confirmed_time", "response_status_code", "comment"]
        ws.append(headers)

        products = MerchPaymentFilter(self.request.GET, queryset=self.get_queryset()).qs
        for payment in products:
            row = []
            for field in headers:
                value = getattr(payment, field)
                if not value:
                    value = ''
                if field in ('id', 'order_id', 'merchant', 'create_at', 'confirmed_time'):
                    row.append(str(value))
                else:
                    row.append(value)
            ws.append(row)
        wb.save(response)
        return response


class MerchantDetail(AuthorRequiredMixin, UpdateView):
    template_name = 'payment/merchant.html'
    model = Merchant
    success_url = reverse_lazy('payment:menu')
    fields = ('name', 'host', 'host_withdraw', 'pay_success_endpoint', 'secret', 'check_balance', 'white_ip',
              'dump_webhook_data', 'merch_viewers')

    def get_context_data(self, **kwargs):
        viewers = self.object.merch_viewers
        print(viewers)
        context = super().get_context_data(**kwargs)
        context['form'] = MerchantForm(instance=self.object)
        viewers_user = User.objects.filter(username__in=viewers)
        context['viewers_user'] = viewers_user
        print(viewers_user)

        return context


class MerchantDelete(AuthorRequiredMixin, SuccessMessageMixin, DeleteView):
    template_name = 'payment/merchant_confirm_delete.html'
    success_url = reverse_lazy('payment:menu')
    model = Merchant
    success_message = 'Delete complete'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.payments.all():
            return HttpResponseBadRequest("You can't delete not empty Merchant. Call to administrator to delete")
        form = self.get_form()
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)


@login_required()
def merchant_test_webhook(request, *args, **kwargs):
    if request.user.role != 'merchant':
        return HttpResponseBadRequest()
    print('merchant_test_webhook')
    order_id = str(uuid.uuid4())
    pk = str(uuid.uuid4())
    merchant_id = (request.POST.get('payment_decline') or
                   request.POST.get('payment_accept') or
                   request.POST.get('withdraw_accept') or
                   request.POST.get('withdraw_decline'))
    if not merchant_id:
        return HttpResponseBadRequest()
    merchant = Merchant.objects.get(pk=merchant_id)
    payment = Payment(merchant=merchant,
                      id=pk,
                      order_id=order_id,
                      amount=random.randrange(10, 3000),
                      create_at=(timezone.now() - datetime.timedelta(minutes=1)),
                      )
    withdraw = Withdraw(
        merchant=merchant,
        id=pk,
        withdraw_id=order_id,
        amount=random.randrange(10, 3000),
        create_at=(timezone.now() - datetime.timedelta(minutes=1)),
    )
    if 'payment_decline' in request.POST:
        payment.status = -1
        data = payment.webhook_data()
        send_payment_webhook.delay(merchant.host, data,
                                   dump_data=payment.merchant.dump_webhook_data)
    elif 'payment_accept' in request.POST:
        payment.confirmed_amount = random.randrange(10, 3000)
        payment.status = 9
        payment.confirmed_time = timezone.now()
        data = payment.webhook_data()
        send_payment_webhook.delay(merchant.host, data, payment.merchant.dump_webhook_data)
    elif 'withdraw_accept' in request.POST:
        withdraw.status = 9
        withdraw.confirmed_time = timezone.now()
        data = withdraw.webhook_data()
        send_withdraw_webhook.delay(merchant.host_withdraw or merchant.host, data,
                                    dump_data=withdraw.merchant.dump_webhook_data)
    else:
        withdraw.status = -1
        data = withdraw.webhook_data()
        send_withdraw_webhook.delay(merchant.host_withdraw or merchant.host, data,
                                    dump_data=withdraw.merchant.dump_webhook_data)

    return JsonResponse(json.dumps(data), safe=False)


@login_required()
def export_payments(request, *args, **kwargs):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="payments.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    headers = ["Id", "order_id", "pay_type", "create_at", 'merchant', "amount", "comission", "confirmed_amount", "status", "user_login", "owner_name", "mask", "referrer", "confirmed_time", "response_status_code", "comment"]
    ws.append(headers)
    if request.user.is_staff:
        products = Payment.objects.all()
    else:
        products = Payment.objects.filter(merchant__owner=request.user)
    for payment in products:
        row = []
        for field in headers:
            value = getattr(payment, field)
            if not value:
                value = ''
            if field in ('id', 'order_id', 'merchant', 'create_at', 'confirmed_time'):
                row.append(str(value))
            else:
                row.append(value)
        ws.append(row)

    # Save the workbook to the HttpResponse
    wb.save(response)
    return response


class MerchStatView(DetailView, ):
    template_name = 'payment/merch_stat.html'
    model = User
    context_object_name = 'merch_user'
    filter = PaymentMerchStatFilter


    # def get_queryset(self):
    #     return Payment.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.object
        payments = Payment.objects.filter(merchant__owner=user).annotate(oper_time=F('confirmed_time') - F('create_at')).all()
        p1 = payments.first()
        t1 = p1.create_at
        t2 = p1.confirmed_time
        print(t2-t1)
        filter = PaymentMerchStatFilter(self.request.GET, queryset=payments)
        context['filter'] = filter
        filtered_payments = filter.qs
        total = filtered_payments.filter()
        confirmed = filtered_payments.filter(status=9)
        declined = filtered_payments.filter(status=-1)

        count_total = total.count()
        total_amount = total.aggregate(total_amount=Sum('amount'))['total_amount']
        confirmed_amount = confirmed.aggregate(confirmed_amount=Sum('amount'))['confirmed_amount']
        count_confirmed = confirmed.count()
        count_declined = declined.count()
        conversion = int(round(count_confirmed / count_total * 100, 0))
        operator_avg_time = filtered_payments.aggregate(operator_avg_time=Avg('oper_time'))['operator_avg_time'].total_seconds()

        context['stat'] = {
            'count_total': count_total,
            'total_amount': total_amount,
            'count_confirmed': count_confirmed,
            'confirmed_amount': confirmed_amount,
            'count_declined': count_declined,
            'conversion': conversion,
            'operator_avg_time': int(operator_avg_time)
        }
        context['stat'] = {
            'count_total': 89952,
            'total_amount': 6431875,
            'count_confirmed': 87338,
            'confirmed_amount': 6245350,
            'count_declined': 2614,
            'conversion': 97,
            'operator_avg_time': 61
        }
        context['balance'] = 84030.07
        return context


class WebhookReceive(APIView):

    def get(self, request, *args, **kwargs):
        logger.debug('WebhookReceive')
        return HttpResponse('ok')

    def post(self, request, *args, **kwargs):
        logger.info('WebhookReceive')
        data = request.data
        logger.info(data)
        return JsonResponse({'status': 'success', 'data': data})


class WebhookRepeat(StaffOnlyPerm, UpdateView, ):
    model = Payment
    success_url = reverse_lazy('payment:payment_list')

    def post(self, request, *args, **kwargs):
        payment = self.get_object()
        data = payment.webhook_data()
        logger.debug(f'Отправка повторного вэбхук {payment.id}: {data}')
        result = send_payment_webhook.delay(url=payment.merchant.host, data=data,
                                                dump_data=payment.merchant.dump_webhook_data)
        return HttpResponse(f'Вэбхук отправлен:<br>{data}')


def on_work(request, *args, **kwargs):
    """Вход-выход на смену"""
    profile = request.user.profile
    oper_id = str(request.user.id)
    operators_on_work: list = SupportOptions.load().operators_on_work

    if not profile.on_work:
        # Если пытаешься включить
        if oper_id not in operators_on_work:
            return HttpResponseBadRequest('Вы не на смене и не можете подключиться')

    if profile.on_work:
        # Если пытаешься выключить
        oper_count = User.objects.filter(profile__on_work=True).count()
        if oper_count <= 1:
            return HttpResponseBadRequest('Вы один на смене и не можете отключиться')
    with transaction.atomic():
        profile.on_work = not profile.on_work
        work = Work(user_id=oper_id, status=profile.on_work)
        work.save()
        profile.save()
        logger.debug(f'Смена смены {request.user}: текущий статус {profile.on_work}')
    return redirect('payment:payment_list')
